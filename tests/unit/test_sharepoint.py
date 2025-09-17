from io import BytesIO
from unittest.mock import patch

from openpyxl import Workbook
import pandas as pd
import pytest

from viadot.exceptions import CredentialError
from viadot.sources import Sharepoint
from viadot.sources.sharepoint import SharepointCredentials


DUMMY_CREDS = {
    "site": "tenant.sharepoint.com",  # pragma: allowlist secret
    "client_id": "dummy_client_id",  # pragma: allowlist secret
    "client_secret": "dummy_client_secret",  # pragma: allowlist secret
    "tenant_id": "dummy_tenant_id",  # pragma: allowlist secret
}
SAMPLE_DF = pd.DataFrame(
    {
        "int_col": [1, 2, 3, 4, 5, None],
        "float_col": [1.1, 2.2, 3.3, 3.0, 5.5, 6.6],
        "str_col": ["a", "b", "c", "d", "e", "f"],
        "nan_col": [None, None, None, None, None, None],
        "mixed_col": [1, "text", None, None, 4.2, "text2"],
    }
)


def create_excel_file():
    """Create a simple Excel file in memory for testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "col_a"
    ws["B1"] = "col_b"
    ws["A2"] = "val1"
    ws["B2"] = "val1"
    ws["A3"] = ""
    ws["B3"] = "val2"
    ws["A4"] = "val2"
    ws["B4"] = "val3"
    ws["A5"] = "NA"
    ws["B5"] = "val4"
    ws["A6"] = "N/A"
    ws["B6"] = "val5"
    ws["A7"] = "#N/A"
    ws["B7"] = "val6"
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


class SharepointMock(Sharepoint):
    def _download_file_stream(self, _url: str | None = None, **kwargs):
        if "nrows" in kwargs:
            msg = "Parameter 'nrows' is not supported."
            raise ValueError(msg)

        return pd.ExcelFile(BytesIO(create_excel_file()))


@pytest.fixture
def sharepoint_mock():
    return SharepointMock(credentials=DUMMY_CREDS)


@pytest.fixture
def sharepoint():
    credentials = {
        "site": "example.sharepoint.com",  # pragma: allowlist secret
        "client_id": "dummy_client_id",  # pragma: allowlist secret
        "client_secret": "dummy_client_secret",  # pragma: allowlist secret
        "tenant_id": "dummy_tenant_id",  # pragma: allowlist secret
    }
    return Sharepoint(credentials=credentials)


def test_valid_credentials():
    credentials = {
        "site": "tenant.sharepoint.com",  # pragma: allowlist secret
        "client_id": "client",  # pragma: allowlist secret
        "client_secret": "secret",  # pragma: allowlist secret
        "tenant_id": "tenant",  # pragma: allowlist secret
    }
    shrp_creds = SharepointCredentials(**credentials)
    assert shrp_creds.site == credentials["site"]
    assert shrp_creds.client_id == credentials["client_id"]
    assert shrp_creds.client_secret == credentials["client_secret"]
    assert shrp_creds.tenant_id == credentials["tenant_id"]


def test_invalid_authentication():
    credentials = {
        "site": "tenant.sharepoint.com",  # pragma: allowlist secret
        "client_id": "client",  # pragma: allowlist secret
        "client_secret": "secret",  # pragma: allowlist secret
        "tenant_id": "tenant",  # pragma: allowlist secret
    }

    s = Sharepoint(credentials=credentials)

    # Patch the GraphClient constructor to simulate an authentication failure
    with patch("viadot.sources.sharepoint.GraphClient") as graph_client:
        graph_client.side_effect = Exception("Authentication failed")

        with pytest.raises(
            CredentialError,
            match="Could not authenticate to tenant.sharepoint.com with provided credentials.",
        ):
            s.get_client()


def test_missing_client_id():
    credentials = {
        "site": "example.sharepoint.com",  # pragma: allowlist secret
        "client_secret": "x",  # pragma: allowlist secret
        "tenant_id": "t",  # pragma: allowlist secret
    }
    with pytest.raises(
        CredentialError,
        match=(
            "'site', 'client_id', 'client_secret' and "
            "'tenant_id' credentials are required."
        ),
    ):
        SharepointCredentials(**credentials)


def test_sharepoint_default_na(sharepoint_mock):
    df = sharepoint_mock.to_df(
        url="test/file.xlsx", na_values=Sharepoint.DEFAULT_NA_VALUES
    )

    assert not df.empty
    assert "NA" not in list(df["col_a"])


def test_sharepoint_custom_na(sharepoint_mock):
    df = sharepoint_mock.to_df(
        url="test/file.xlsx",
        na_values=[v for v in Sharepoint.DEFAULT_NA_VALUES if v != "NA"],
    )

    assert not df.empty
    assert "NA" in list(df["col_a"])


def test__get_file_extension(sharepoint_mock):
    url_excel = "https://tenant.sharepoint.com/sites/site/file.xlsx"
    url_dir = "https://tenant.sharepoint.com/sites/site/"
    url_txt = "https://tenant.sharepoint.com/sites/site/file.txt"

    excel_ext = sharepoint_mock._get_file_extension(url=url_excel)
    txt_ext = sharepoint_mock._get_file_extension(url=url_txt)
    dir_ext = sharepoint_mock._get_file_extension(url=url_dir)

    assert excel_ext == ".xlsx"
    assert txt_ext == ".txt"
    assert dir_ext == ""


def test__is_file(sharepoint_mock):
    is_file = sharepoint_mock._is_file(url="https://example.com/file.xlsx")
    assert is_file is True

    is_file = sharepoint_mock._is_file(url="https://example.com/dir")
    assert is_file is False


def test__parse_excel_single_sheet(sharepoint_mock):
    excel_file = sharepoint_mock._download_file_stream()
    result_df = sharepoint_mock._parse_excel(excel_file, sheet_name="Sheet1")
    expected = pd.DataFrame(
        {
            "col_a": ["val1", "", "val2", "NA", "N/A", "#N/A"],
            "col_b": ["val1", "val2", "val3", "val4", "val5", "val6"],
        }
    )

    assert result_df["col_b"].equals(expected["col_b"])


def test__parse_excel_string_dtypes(sharepoint_mock):
    excel_file = sharepoint_mock._download_file_stream()
    result_df = sharepoint_mock._parse_excel(excel_file, sheet_name="Sheet1")

    for column in result_df.columns:
        assert result_df[column].dtype == object


def test__load_and_parse_not_valid_extension(sharepoint_mock):
    with pytest.raises(ValueError):  # noqa: PT011
        sharepoint_mock._load_and_parse(file_url="https://example.com/file.txt")


def test_scan_sharepoint_folder_valid_url(sharepoint_mock):
    url = "https://company.sharepoint.com/sites/site_name/final_folder/"

    # Build a fake Graph client hierarchy to return children with names
    class _FakeChildren:
        def get(self):
            class _Exec:
                def execute_query(self):
                    class _Item:
                        def __init__(self, name):
                            self.name = name

                    return [_Item("file1.txt"), _Item("file2.txt")]

            return _Exec()

    class _FakeDriveItem:
        def get(self):
            class _Exec:
                def execute_query(self):
                    # Return an object that has .children
                    class _FolderItem:
                        children = _FakeChildren()

                    return _FolderItem()

            return _Exec()

    class _FakeByUrl:
        @property
        def drive_item(self):
            return _FakeDriveItem()

    class _FakeShares:
        def by_url(self, _url):
            return _FakeByUrl()

    class _FakeClient:
        shares = _FakeShares()

    with patch.object(Sharepoint, "get_client", return_value=_FakeClient()):
        expected_files = [
            (
                "https://company.sharepoint.com/sites/site_name/"
                "final_folder/file1.txt"
            ),
            (
                "https://company.sharepoint.com/sites/site_name/"
                "final_folder/file2.txt"
            ),
        ]

        result = sharepoint_mock.scan_sharepoint_folder(url)
        assert result == expected_files


def test_scan_sharepoint_folder_invalid_url(sharepoint_mock):
    url = "https://company.sharepoint.com/folder/sub_folder/final_folder"

    with pytest.raises(ValueError, match="URL does not contain '/sites/' segment."):
        sharepoint_mock.scan_sharepoint_folder(url)


def test_scan_sharepoint_folder_empty_response(sharepoint_mock):
    url = (
        "https://company.sharepoint.com/sites/site_name/folder/sub_folder/final_folder"
    )

    # Fake client that returns no children
    class _FakeChildren:
        def get(self):
            class _Exec:
                def execute_query(self):
                    return []

            return _Exec()

    class _FakeDriveItem:
        def get(self):
            class _Exec:
                def execute_query(self):
                    class _FolderItem:
                        children = _FakeChildren()

                    return _FolderItem()

            return _Exec()

    class _FakeByUrl:
        @property
        def drive_item(self):
            return _FakeDriveItem()

    class _FakeShares:
        def by_url(self, _url):
            return _FakeByUrl()

    class _FakeClient:
        shares = _FakeShares()

    with patch.object(Sharepoint, "get_client", return_value=_FakeClient()):
        result = sharepoint_mock.scan_sharepoint_folder(url)
        assert result == []


def test_download_file_stream_unsupported_param(sharepoint_mock):
    url = "https://company.sharepoint.com/sites/site_name/folder/test_file.xlsx"

    with pytest.raises(ValueError, match="Parameter 'nrows' is not supported."):
        sharepoint_mock._download_file_stream(url, nrows=10)


def test_successful_download(sharepoint):
    url = "https://example.sharepoint.com/sites/site/Shared%20Documents/file.xlsx"
    with patch.object(
        Sharepoint,
        "_download_file_stream",
        return_value=pd.ExcelFile(BytesIO(create_excel_file())),
    ):
        df = sharepoint.to_df(url)
        assert not df.empty
        assert len(df) == 6


def test_access_denied(sharepoint):
    url = (
        "https://example.sharepoint.com/sites/site/"
        "Shared%20Documents/restricted_file.xlsx"
    )
    with (
        patch.object(
            Sharepoint,
            "_download_file_stream",
            side_effect=PermissionError("403 Forbidden"),
        ),
        pytest.raises(PermissionError),
    ):
        sharepoint.to_df(url)


def test_invalid_excel_file(sharepoint):
    url = "https://example.sharepoint.com/sites/site/Shared%20Documents/file.xlsx"
    with (
        patch.object(
            Sharepoint,
            "_download_file_stream",
            side_effect=ValueError("Excel file format cannot be determined"),
        ),
        pytest.raises(ValueError, match="Excel file format cannot be determined"),
    ):
        sharepoint.to_df(url)


def test_mixed_file_extensions(sharepoint):
    url = "https://example.sharepoint.com/sites/site/Shared%20Documents/folder"
    with (
        patch.object(Sharepoint, "scan_sharepoint_folder") as mock_scan,
        patch.object(Sharepoint, "_load_and_parse") as mock_load_and_parse,
    ):
        mock_scan.return_value = [
            url + "/file1.xlsx",
            url + "/file2.txt",
            url + "/file3.pdf",
        ]
        mock_load_and_parse.return_value = pd.DataFrame(
            {
                "col_a": ["val1", "", "val2", "NA", "N/A", "#N/A"],
                "col_b": ["val1", "val2", "val3", "val4", "val5", "val6"],
            }
        )
        df = sharepoint.to_df(url)

        # Verify that _load_and_parse was only called for .xlsx files
        mock_load_and_parse.assert_called_once_with(
            file_url=url + "/file1.xlsx",
            sheet_name=None,
            na_values=None,
            **{},
        )

        assert not df.empty
        assert len(df) == 6


def test_empty_folder(sharepoint):
    url = "https://example.sharepoint.com/sites/site/Shared%20Documents/folder"
    with (
        patch.object(Sharepoint, "scan_sharepoint_folder") as mock_scan,
    ):
        mock_scan.return_value = []
        df = sharepoint.to_df(url)
        assert df.empty
